# -*- coding: utf-8 -*-
"""最终装配: 读workflow结果(descs/funds) → 三张sheet → 桌面 Track/美股选股追踪.xlsx"""
import json, sys, warnings, datetime, re
import pandas as pd, numpy as np, openpyxl, yfinance as yf
from openpyxl.styles import Font, Alignment, PatternFill
warnings.filterwarnings('ignore')
exec(open('build_xlsx.py').read().split("# ══ Sheet 1")[0])   # 复用样式/数据装载 (首段无残注释)
wf=json.load(open(sys.argv[1]))                                # workflow返回 {descs, funds}
DESC={d['t']:d['desc'] for d in wf.get('descs',[]) if d.get('t')}
def brief(t): return DESC.get(t,'')

# ── 13F: issuer名 → ticker (用宇宙名模糊匹配) ──
uni_list=json.load(open('universe.json'))
ABBR={'FINL':'FINANCIAL','COS':'COMPANIES','CO':'COMPANY','PETE':'PETROLEUM','HLDGS':'HOLDINGS','HLDG':'HOLDING','TECHNOLOGIES':'TECH','TECHNOLOGY':'TECH',
      'INTL':'INTERNATIONAL','SVCS':'SERVICES','SVC':'SERVICE','GRP':'GROUP','MFG':'MANUFACTURING','SYS':'SYSTEMS','PHARMA':'PHARMACEUTICALS',
      'PHARMACEUTICAL':'PHARMACEUTICALS','LABS':'LABORATORIES','ELEC':'ELECTRIC','ENTMT':'ENTERTAINMENT','COMMUNICATIONS':'COMM',
      'INDS':'INDUSTRIES','IND':'INDUSTRIES','RES':'RESOURCES','NATL':'NATIONAL','AMERN':'AMERICAN','SEMICONDUCTOR':'SEMI','BANCORP':'BANK','BREWING':'BEVERAGE','INSTRS':'INSTRUMENTS','MTR':'MOTOR','MTRS':'MOTORS','SOUTHN':'SOUTHERN','MATLS':'MATERIALS','FGHT':'FREIGHT','PPTYS':'PROPERTIES','PPTY':'PROPERTY','RESH':'RESEARCH','THERAPEUTIC':'THERAPEUTICS','BIOSCIENCE':'BIOSCIENCES','MED':'MEDICAL','ENTERPRISES':'ENTERPRISE','BK':'BANK','HOLDLINGS':'HOLDINGS','BANCSHARES':'BANK','BANCORPORATION':'BANK','BLDG':'BUILDING','PRODS':'PRODUCTS','FMRS':'FARMERS','MKT':'MARKET','MKTS':'MARKETS'}
DROP={'INC','CORP','CO','COMPANY','COMPANIES','CORPORATION','INCORPORATED','LTD','PLC','LLC','SA','NV','AG','NEW','CLASS','CL','A','B','C','COM','COMMON','STK','STOCK','SHS','ORD','ADR','ADS','DEL','THE','HOLDINGS','HOLDING','GROUP','TECH','TRUST','FD','ETF'}
import re, difflib
def norm(s):
    s=(s or '').upper()
    s=re.sub(r'\(.*?\)','',s)
    s=re.sub(r"['.]",'',s)
    s=re.sub(r'\bP\s*L\s*C\b','PLC',s); s=re.sub(r'\bN\s*V\b','NV',s); s=re.sub(r'\bS\s*A\b','SA',s)
    s=re.sub(r'\b(AMERICAN DEPOSITARY|DEPOSITARY|SUBORDINATE VOTING|VOTING SHARES|NEW YORK REGISTRY|REGISTRY SHARES|COMMON STOCK|ORDINARY SHARES|CAPITAL STOCK|EACH REPRESENTING|WARRANTS?|UNITS?|SPONSORED|REPRESENTING)\b.*$','',s)
    w=re.sub(r'[^A-Z0-9 ]',' ',s).split()
    w=[ABBR.get(x,x) for x in w]
    w=[x for x in w if x not in DROP]
    return ''.join(w)
name2t={}; keys=[]
for u in uni_list:
    k=norm(u['name'])
    if k and k not in name2t: name2t[k]=u['t']; keys.append(k)
def issuer2t(issuer):
    k=norm(issuer)
    if not k: return None
    if k in name2t: return name2t[k]
    # 前缀: 发行人名常被截断; 短的一方至少要占长的60%(防 COCACOLA-EUROPACIFIC 落到 KO)
    for nk in keys:
        a,b=(k,nk) if len(k)<=len(nk) else (nk,k)
        if len(a)>=6 and b.startswith(a) and len(a)/len(b)>=0.5: return name2t[nk]
    # 模糊: 只对较长的名字, 且前3字母必须相同(防 EROCK→MERCK / AMBAC→UMB / FERVO→DEVON)
    if len(k)>=8:
        m=difflib.get_close_matches(k,keys,n=1,cutoff=0.88)
        if m and m[0][:3]==k[:3]: return name2t[m[0]]
    return None
SECT={u['t']:u.get('sector') for u in uni_list}

# 按股票聚合: 谁新建/谁加仓
agg={}
for f in wf.get('funds',[]):
    if not f.get('found'): continue
    fn=f['name']; per=f.get('period','')
    for kind,lst in (('新建仓',f.get('new',[])),('加仓',f.get('add',[]))):
        for h in lst:
            t=issuer2t(h.get('issuer',''))
            if not t: continue
            a=agg.setdefault(t,dict(t=t,issuer=h['issuer'],new=[],add=[],value=0,period=per))
            tag = fn if kind=='新建仓' else f"{fn}(+{h.get('chg_pct',0):.0f}%)"
            a['new' if kind=='新建仓' else 'add'].append(tag)
            a['value']+=h.get('value_k',0)
print(f"13F聚合: {len(agg)} 只股票被新建/加仓 (来自{sum(1 for f in wf.get('funds',[]) if f.get('found'))}家基金)")

# 补这批股票的基本面(不在534里的)
need=[t for t in agg if t not in fund]
if need:
    import concurrent.futures as cf
    def one(t):
        try:
            i=yf.Ticker(t).info; rev=i.get('totalRevenue'); ni=i.get('netIncomeToCommon')
            return t,dict(px=i.get('currentPrice') or i.get('regularMarketPrice'),mc=i.get('marketCap'),
                          rev_m=round(rev/1e6,1) if rev else None,npm=round(ni/rev*100,2) if (rev and ni) else None,
                          sector=i.get('sector'),name=i.get('longName') or t)
        except Exception: return t,{}
    with cf.ThreadPoolExecutor(max_workers=8) as ex:
        for t,d in ex.map(one,need): fund[t]=d
    print(f"补基本面 {len(need)} 只")
# 1月涨跌用价格矩阵(不在矩阵里的拉一下)
def m1now(t):
    try:
        if t in px.columns: s=px[t].dropna()
        else: s=yf.Ticker(t).history(period='2mo')['Close'].dropna()
        return round((float(s.iloc[-1])/float(s.iloc[-22])-1)*100,2) if len(s)>=22 else None
    except Exception: return None

# ══ 建簿 ══
wb=openpyxl.Workbook(); wb.remove(wb.active)
exec(open('build_xlsx.py').read().split("# ══ Sheet 1")[1].split("# ══ Sheet 3")[0].split("\n",1)[1])  # Sheet1 (丢首行残注释)

# ══ Sheet 2: 基金新增/加仓 ══
ws2=wb.create_sheet('2_基金新建加仓')
widths(ws2,{'A':9,'B':26,'C':16,'D':10,'E':10,'F':11,'G':11,'H':11,'I':44,'J':44,'K':46})
put(ws2,1,1,f'Long-only/Long-bias 基金 · 最近一期13F 新建仓与加仓≥20% · 更新至 {TODAY}',bold=True)
nf=sum(1 for f in wf.get('funds',[]) if f.get('found'))
put(ws2,2,1,f'{nf}家基金(SEC EDGAR 13F-HR, 季度滞后45天) | 同一股被多家新建/加仓=共识信号 | 按被提及基金数排序')
put(ws2,2,7,'加仓门槛'); put(ws2,2,8,0.20,'pct',fill=YEL)
C2=['代码','公司','板块','现价$','市值$B','近1月%','净利率%','年营收$M','新建仓基金','加仓基金','这家做什么']
header(ws2,4,C2,rightfrom=3)
r=5
for t,a in sorted(agg.items(),key=lambda kv:-(len(kv[1]['new'])+len(kv[1]['add']))):
    f=fund.get(t,{})
    put(ws2,r,1,t); put(ws2,r,2,(f.get('name') or a['issuer'])[:40]); put(ws2,r,3,f.get('sector') or SECT.get(t) or '')
    put(ws2,r,4,f.get('px'),'usd'); put(ws2,r,5,round(f['mc']/1e9,1) if f.get('mc') else None,'1dp')
    m=m1now(t); put(ws2,r,6,m/100 if m is not None else None,'pct1')
    put(ws2,r,7,f.get('npm')/100 if f.get('npm') is not None else None,'pct1')
    put(ws2,r,8,f.get('rev_m'),'int')
    put(ws2,r,9,'; '.join(a['new']),wrap=True); put(ws2,r,10,'; '.join(a['add']),wrap=True)
    put(ws2,r,11,brief(t),wrap=True); r+=1
ws2.freeze_panes='A5'; ws2.auto_filter.ref=f'A4:K{max(r-1,5)}'
print(f"Sheet2: {r-5} 行")

exec(open('build_xlsx.py').read().split("# ══ Sheet 3")[1].split("\n",1)[1])   # Sheet3 (丢首行残注释)
# 表顺序: 1,2,3
wb._sheets=[wb['1_跳涨10%以上'],wb['2_基金新建加仓'],wb['3_沉寂后启动板块']]
for w in wb:
    for row in w.iter_rows():
        for c in row:
            if c.value is not None and c.font.name!=FONT: c.font=F(b=bool(c.font.bold))
out=sys.argv[2] if len(sys.argv)>2 else '/Users/huaichuaibeimeng/Desktop/Track/美股选股追踪.xlsx'
wb.save(out); print(f"✓ 保存 {out}")
