# -*- coding: utf-8 -*-
"""选股文件夹 材料清单 + 2024至今股价 (Buwen 9/3 10:35 令)"""
import json,os,datetime,warnings,csv,re
import pandas as pd,numpy as np,openpyxl
from openpyxl.styles import Font,Alignment,PatternFill
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')
FONT="等线"
def F(b=False,sz=11): return Font(name=FONT,size=sz,bold=b)
FMT={'int':'_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)','2dp':'_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)','text':'@','date':'yyyy-mm-dd'}
def put(ws,r,c,v,key='text',bold=False,wrap=False):
    x=ws.cell(r,c,v); x.font=F(b=bold); x.number_format=FMT[key]
    x.alignment=Alignment(horizontal="right" if key in('int','2dp') else "left",wrap_text=wrap,vertical="top"); return x
def header(ws,row,cols,rightfrom=99):
    for i,t in enumerate(cols,1):
        c=ws.cell(row,i,t); c.font=F(b=True); c.alignment=Alignment(horizontal="right" if i>rightfrom else "left",vertical="bottom",wrap_text=True)
BASE=os.path.expanduser('~/Desktop/选股'); EC=f'{BASE}/earnings call'
ov=json.load(open('overview.json')); mm=json.load(open('models_meta.json')); rm=json.load(open('reports_map.json'))
cov={r['ticker']:r for r in csv.DictReader(open(f'{EC}/_原始与索引/coverage.csv'))}
ecdirs={d for d in os.listdir(EC) if not d.startswith(('_','.'))}
def calls(t):
    if t not in ecdirs: return 0,'',''
    fs=sorted(f[:-4] for f in os.listdir(f'{EC}/{t}') if f.endswith('.txt'))
    return len(fs),(fs[0] if fs else ''),(fs[-1] if fs else '')
tick=sorted(set(ov)|ecdirs|set(mm)|{k for k in rm['by_ticker'] if re.match(r'^[A-Z]',k)},key=lambda t:(ov.get(t,{}).get('cat') or '零 未分类',t))
tick=[t for t in tick if not t.startswith('_')]
wb=openpyxl.Workbook(); ws=wb.active; ws.title='1_材料清单'
TODAY=datetime.date.today().isoformat()
put(ws,1,1,f'选股文件夹 材料清单 · {len(tick)}只 · 三类材料: model(GS模型, 实体在 Semi Research/0_Models) / 卖方研报(高盛PDF, 研报/) / earnings call(逐季全文, earnings call/) · 生成 {TODAY}',bold=True)
put(ws,2,1,'备注: "缺什么"列汇总三类缺口; 覆盖总览里标❌的股票=只在清单上没有任何材料; Call期数来自实际txt文件数; 研报按文件名括号里的ticker归类, 33份行业/主题研报不归个股(见Sheet3)')
cols=['类别','代码','公司/说明','Model','Model文件','Model sheet数','Model年份覆盖','Model更新日','研报数','研报最新标题','Call期数','Call最早','Call最新','FY2021起期数','缺什么','覆盖总览备注']
header(ws,4,cols,rightfrom=5)
r=5
for t in tick:
    o=ov.get(t,{}); m=mm.get(t); reps=[x for x in rm['by_ticker'].get(t,[]) if not x['dup']]; n,c0,c1=calls(t)
    miss=[]
    if not m: miss.append('无model')
    if not reps: miss.append('无研报')
    if n==0: miss.append('无call')
    elif n<8: miss.append(f'call仅{n}期')
    put(ws,r,1,o.get('cat') or '零 未分类'); put(ws,r,2,t); put(ws,r,3,o.get('desc') or '')
    put(ws,r,4,'有' if m else '无'); put(ws,r,5,os.path.basename(m['path']) if m else '')
    put(ws,r,6,m['n_sheets'] if m else None,'int'); put(ws,r,7,(f"{m['yr_min']}-{m['yr_max']}" if m and m.get('yr_min') else '') if m else '')
    put(ws,r,8,m['mtime'] if m else ''); put(ws,r,9,len(reps),'int')
    put(ws,r,10,max(reps,key=lambda x:x['size'])['file'][:-4] if reps else '')
    put(ws,r,11,n,'int'); put(ws,r,12,c0); put(ws,r,13,c1); put(ws,r,14,int(cov[t]['FY2021起']) if t in cov else None,'int')
    put(ws,r,15,'、'.join(miss) if miss else '齐'); put(ws,r,16,o.get('note') or '')
    r+=1
for k,w in {'A':16,'B':9,'C':26,'D':6,'E':20,'F':8,'G':11,'H':11,'I':7,'J':60,'K':8,'L':10,'M':10,'N':9,'O':16,'P':28}.items(): ws.column_dimensions[k].width=w
ws.freeze_panes='C5'
# Sheet2 股价
px=pd.read_pickle('px_2024.pkl'); ymap=json.load(open('px_2024_meta.json'))['ymap']; inv={v:k for k,v in ymap.items()}
px=px[[c for c in px.columns if c!='.DS_Store']].dropna(how='all')
ws2=wb.create_sheet('2_股价2024至今')
put(ws2,1,1,f'收盘价 2024-01-02 至 {px.index[-1].date()} · 来源 Yahoo Finance(不复权Close) · {px.shape[1]}只 · SKHynix=000660.KS / Samsung=005930.KS 为本币(韩元), 其余美元',bold=True)
put(ws2,2,1,'备注: 空格=当日无数据(未上市/停牌/Yahoo缺); 最后一行若为今天且部分为空属正常(美股未收盘/韩股已交易)')
ordered=[t for t in tick if ymap.get(t,t) in px.columns]
header(ws2,4,['日期']+ordered,rightfrom=1)
for i,(d,row) in enumerate(px.iterrows()):
    put(ws2,5+i,1,d.to_pydatetime(),'date')
    for j,t in enumerate(ordered):
        v=row.get(ymap.get(t,t))
        if v==v: put(ws2,5+i,2+j,round(float(v),2),'2dp')
ws2.column_dimensions['A'].width=12
for j in range(len(ordered)): ws2.column_dimensions[get_column_letter(2+j)].width=9
ws2.freeze_panes='B5'
# Sheet3 行业/主题研报
ws3=wb.create_sheet('3_行业主题研报'); put(ws3,1,1,'不归个股的行业/主题研报(高盛) · 文件在 研报/',bold=True); header(ws3,4,['文件名','大小KB'],rightfrom=1)
for i,f in enumerate(sorted(rm['sector'])):
    put(ws3,5+i,1,f[:-4]); put(ws3,5+i,2,round(os.path.getsize(f'{BASE}/研报/{f}')/1024),'int')
ws3.column_dimensions['A'].width=120; ws3.column_dimensions['B'].width=9
out=f'{BASE}/选股_材料清单与股价.xlsx'; wb.save(out); print('✓',out,'| 清单',len(tick),'只 | 股价',px.shape,'| 缺口统计:')
import collections; print(collections.Counter(ws.cell(i,15).value for i in range(5,r)))
