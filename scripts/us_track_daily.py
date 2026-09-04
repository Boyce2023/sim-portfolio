#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""美股选股追踪.xlsx 每日更新 (2026-09-02 Buwen令'每日更新, 一直加新行, 最新的放最上面')
Sheet1: 扫昨日≥10%跳涨 → 插到第5行(最新在上), 不动已有行(Buwen会手改)
Sheet3: 只刷数值列, 不动结构
Sheet2: 13F季度更新, 本脚本不动
⛔覆盖守卫: 表头与预期不符则中止(防冲掉Buwen的改动)。用法: us_track_daily.py [--dry OUT.xlsx]
"""
import sys, os, json, datetime, warnings
import pandas as pd, numpy as np, yfinance as yf, openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def nasdaq_jumps(min_pct=0.10, min_mc_b=5.0):
    """当日跳涨主源: NASDAQ screener(收盘即可得, 不依赖Yahoo日线落盘)。
    返回 [(ticker, name, pct, last, mc_b)]; 失败返回 None 让调用方回退 Yahoo。"""
    import urllib.request, json as _j, time as _t
    rows=None
    for attempt in range(4):
        try:
            req=urllib.request.Request(
                "https://api.nasdaq.com/api/screener/stocks?tableonly=true&limit=6000&download=true",
                headers={"User-Agent":"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
                         "Accept":"application/json","Accept-Language":"en-US,en;q=0.9"})
            rows=_j.load(urllib.request.urlopen(req,timeout=60))['data']['rows']
            break
        except Exception as e:
            if attempt==3:
                print(f"[NASDAQ 4次重试均失败: {str(e)[:60]}]")
                return None
            _t.sleep(20*(attempt+1))
    if rows is None: return None
    out=[]
    for r in rows:
        try:
            pct=float(str(r['pctchange']).replace('%',''))/100.0
            mc=float(r['marketCap'] or 0)/1e9
            last=float(str(r['lastsale']).replace('$',''))
        except Exception:
            continue
        if pct>=min_pct and mc>=min_mc_b:
            out.append((r['symbol'], r['name'][:40], pct, last, round(mc,1)))
    out.sort(key=lambda x:-x[2])
    return out


warnings.filterwarnings('ignore')
SRC=os.path.expanduser('~/Desktop/Track/美股选股追踪.xlsx')
UNI=os.path.expanduser('~/claude-projects/sim-portfolio/data/us_track_universe.json')
OUT=sys.argv[sys.argv.index('--dry')+1] if '--dry' in sys.argv else SRC
FONT="等线"
def F(b=False): return Font(name=FONT,size=11,bold=b)
FMT={'int':'_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)','1dp':'_(* #,##0.0_);_(* \\(#,##0.0\\);_(* "-"??_);_(@_)',
     'pct1':'0.0%','usd':'_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)','text':'@'}
H1=['跳涨日','代码','公司','板块','前日收盘$','跳涨日收盘$','当日涨跌%','现价$','市值$B','此前1月%','此后1月%','这家做什么']
H3=['ETF','板块','现价$','24月%','12月%','3月%','1月%','12月超额pp','24月超额pp','命中','说明']

def put(ws,r,c,v,key='text',wrap=False):
    x=ws.cell(r,c,v); x.font=F(); x.number_format=FMT[key]
    x.alignment=Alignment(horizontal="right" if key!='text' else "left",wrap_text=wrap,vertical="top")

def main():
    if not os.path.exists(SRC): sys.exit(f"⛔ 找不到 {SRC}")
    wb=openpyxl.load_workbook(SRC)
    ws1=wb['1_跳涨10%以上']; ws3=wb['3_沉寂后启动板块']
    for ws,H in ((ws1,H1),(ws3,H3)):
        cur=[ws.cell(4,c).value for c in range(1,len(H)+1)]
        if cur!=H: sys.exit(f"⛔ 中止: {ws.title} 表头与预期不符(可能被手改), 现有{cur[:5]}... 预期{H[:5]}...")
    uni=json.load(open(UNI)); tk=[x['t'] for x in uni]; meta={x['t']:x for x in uni}
    spy=yf.Ticker('SPY').history(period='10d')['Close'].dropna()
    days=[pd.Timestamp(d.date()) for d in spy.index]
    last=days[-1]; prev=days[-2]
    # ⛔2026-09-04修: Yahoo日线常在收盘后数小时才落盘, 会漏掉整个交易日(9/3漏了13只)。
    # 改用 NASDAQ screener 作当日跳涨主源(收盘即可得), Yahoo 仅用于补充历史字段。
    nd=nasdaq_jumps()
    if nd is not None:
        import datetime as _dt, zoneinfo
        # NASDAQ 给的是"最近一个已收盘交易日"的涨跌幅。⛔不能用 Yahoo 日线的 last(它可能滞后一天),
        # 改用美东时间推算: 已过16:00且为工作日→今天; 否则回退到上一个工作日。
        et=_dt.datetime.now(zoneinfo.ZoneInfo('America/New_York'))
        d0=et.date()
        if et.hour < 16 or et.weekday() > 4:
            d0 = d0 - _dt.timedelta(days=1)
        while d0.weekday() > 4:
            d0 = d0 - _dt.timedelta(days=1)
        last_used=pd.Timestamp(d0)
        if last_used != last:
            print(f"[日期口径] NASDAQ对应交易日={last_used.date()}, Yahoo日线最新={last.date()} (Yahoo滞后, 以NASDAQ为准)")
        have=set()
        for r in range(5,ws1.max_row+1):
            if str(ws1.cell(r,1).value)==str(last_used.date()): have.add(ws1.cell(r,2).value)
        new=[]
        for sym,name,pct,lastpx,mcb in nd:
            if sym in have: continue
            new.append(dict(t=sym,name=name,prev=round(lastpx/(1+pct),2),close=round(lastpx,2),chg=pct,before=None,mc=mcb,src='nasdaq'))
        print(f"[主源NASDAQ] {last_used.date()} 跳涨≥10%且市值≥50亿: {len(nd)}只, 其中新增 {len(new)}只")
    else:
        print("[主源NASDAQ失败, 回退Yahoo日线]")
        last_used=last
        have=set()
        for r in range(5,ws1.max_row+1):
            if str(ws1.cell(r,1).value)==str(last.date()): have.add(ws1.cell(r,2).value)
        px=yf.download(tk,period='2mo',progress=False,auto_adjust=False,threads=True)['Close']
        px.index=pd.to_datetime([d.date() for d in px.index])
        new=[]
        for t in tk:
            if t not in px.columns or t in have: continue
            s=px[t].dropna()
            if last not in s.index or prev not in s.index: continue
            chg=float(s.loc[last])/float(s.loc[prev])-1
            if chg>=0.10:
                before=(float(s.loc[last])/float(s.iloc[max(0,len(s[s.index<=last])-22)])-1) if len(s)>=22 else None
                new.append(dict(t=t,name=None,prev=round(float(s.loc[prev]),2),close=round(float(s.loc[last]),2),chg=chg,before=before,mc=None,src='yahoo'))
    new.sort(key=lambda x:-x['chg'])
    if new:
        ws1.insert_rows(5,amount=len(new))
        for i,j in enumerate(new):
            r=5+i; t=j['t']; m=meta.get(t,{})
            try: f=yf.Ticker(t).info
            except Exception: f={}
            put(ws1,r,1,str(last_used.date())); put(ws1,r,2,t); put(ws1,r,3,(f.get('longName') or j.get('name') or m.get('name') or t)[:40])
            put(ws1,r,4,f.get('sector') or m.get('sector') or ''); put(ws1,r,5,j['prev'],'usd'); put(ws1,r,6,j['close'],'usd')
            put(ws1,r,7,j['chg'],'pct1'); put(ws1,r,8,f.get('currentPrice') or j['close'],'usd')
            put(ws1,r,9,round(f['marketCap']/1e9,1) if f.get('marketCap') else (j.get('mc') or m.get('mc_b')),'1dp')
            put(ws1,r,10,j['before'],'pct1'); put(ws1,r,11,None,'pct1')
            put(ws1,r,12,'(新增, 待写介绍)',wrap=True)
        ws1.auto_filter.ref=f'A4:L{ws1.max_row}'
    # ── Sheet3: 刷数值 ──
    spy3=yf.Ticker('SPY').history(period='3y')['Close'].dropna()
    sr=lambda n:(float(spy3.iloc[-1])/float(spy3.iloc[-n])-1)*100
    s24,s12,s1=sr(505),sr(253),sr(22)
    for r in range(5,ws3.max_row+1):
        etf=ws3.cell(r,1).value
        if not etf: continue
        try:
            s=yf.Ticker(etf).history(period='3y')['Close'].dropna()
            if len(s)<253: continue
            now=float(s.iloc[-1]); r24=(now/float(s.iloc[-505])-1)*100 if len(s)>=505 else None
            r12=(now/float(s.iloc[-253])-1)*100; r3=(now/float(s.iloc[-64])-1)*100; r1=(now/float(s.iloc[-22])-1)*100
            u12=r12-s12; u24=(r24-s24) if r24 is not None else None; hit=u12<0 and (r1-s1)>5
            put(ws3,r,3,round(now,2),'usd'); put(ws3,r,4,r24/100 if r24 is not None else None,'pct1')
            put(ws3,r,5,r12/100,'pct1'); put(ws3,r,6,r3/100,'pct1'); put(ws3,r,7,r1/100,'pct1')
            put(ws3,r,8,round(u12,1),'1dp'); put(ws3,r,9,round(u24,1) if u24 is not None else None,'1dp')
            put(ws3,r,10,'⭐命中' if hit else '')
        except Exception: pass
    t1=ws1.cell(1,1).value or ''; ws1.cell(1,1).value=t1.split('· 更新至')[0].rstrip()+f'· 更新至 {last_used.date()}'
    t3=ws3.cell(1,1).value or ''; ws3.cell(1,1).value=t3.split('· 更新至')[0].rstrip()+f'· 更新至 {last_used.date()}'
    wb.save(OUT)
    print(f"✓ {OUT} | 交易日{last_used.date()} 新增跳涨{len(new)}条: {[j['t'] for j in new][:15]} | Sheet3已刷")

if __name__=='__main__': main()
