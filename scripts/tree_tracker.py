#!/usr/bin/env python3
"""
产品树日度涨跌追踪器 (2026-08-04建, Buwen要求)
────────────────────────────────────────────────
产出: ~/Desktop/track/产品树涨跌追踪.xlsx
  Sheet1「近30日」   : 产品树第一列, B/C/D=近30/10/3日累计(⭐Excel公式非硬编码), 后接30个日期列。每天刷新
  Sheet2「历史累计」 : 产品树第一行(列头), 日期第一列往下。每天最下面追加, ⛔绝不覆盖历史(默认灌6个月)

格式(xlsx skill标准):
  · Arial 9pt / 无斑马纹 / 无彩色tab
  · 蓝色(0,0,255)=硬编码输入(日度涨跌数据) / 黑色=公式(累计列)
  · 负数用括号 (1.5) 不是 -1.5
  · 累计列用 PRODUCT 公式,Excel自己算,可随数据更新

用法: python3 tree_tracker.py                    # 更新到今天(历史sheet首建灌6个月)
     python3 tree_tracker.py --history-days 180  # 指定历史深度
"""
import sys, os, argparse, datetime as dt
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('NO_PROXY', '*')
import ifind_data_layer as ifd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

OUT_DIR = os.path.expanduser('~/Desktop/track')
OUT = os.path.join(OUT_DIR, '产品树涨跌追踪.xlsx')

TREES = {
 'AI算力(VR200)':      ['300308','002463','601138','300502'],
 'AI端侧(手机换机)':    ['002475','603893','603986','300458'],
 '半导体设备国产化':     ['002371','688012','688072','300604'],
 '存储扩产':           ['603986','301308','688525','000021'],
 '人形机器人':          ['688017','002472','300748','603662'],
 '智能驾驶Robotaxi':   ['002920','603596','300627','300496'],
 '苹果新形态':          ['002273','600552','300709','002415'],
 '固态电池':           ['002812','300568','002460','603659'],
 '电动车':             ['002594','300750','002460','600733'],
 '创新药MNC扫货':       ['600276','688506','688131','300347'],
 'CXO/CDMO':          ['603259','002821','300759','300363'],
 'GLP-1减肥降糖':       ['000963','300199','688076','301393'],
 '医疗器械IVD':         ['300760','688617','300003','002223'],
 '中药院内制剂':         ['600085','000538','600436','600332'],
 '脑机/手术机器人':      ['688017','688050','300633','688029'],
 'AI供电(电网扩容)':     ['600089','002028','600312','601179'],
 '电力核电绿电特高压':    ['601611','601985','600886','000400'],
 '可控核聚变':          ['300185','002080','601106','300772'],
 '商业航天':            ['600118','002025','688055','300101'],
 '军工航发/主战装备':     ['600760','600893','300034','002179'],
 'eVTOL低空经济':       ['002526','300424','600038','002013'],
 '战略小金属管制':        ['600111','300748','000657','600549'],
 '稀土永磁':            ['600111','300748','600259','000831'],
 '钨硬质合金':           ['000657','600549','002378','688059'],
 '制冷剂氟化工':          ['600160','002407','603379','603505'],
 '高端白酒':             ['600519','000858','000568','600779'],
 '情绪经济/新消费':        ['603517','002458','300896','603605'],
 '免税出行链':            ['601888','600754','601111','600115'],
 '家电以旧换新':           ['000333','000651','688169','002032'],
 '券商金融IT':            ['600030','000776','300033','603383'],
 '农业种业转基因':          ['002041','000998','600598','002385'],
 '猪周期':                ['300498','002714','000876','002124'],
 '品牌中药OTC保健':        ['600436','600085','000538','600993'],
}
sfx = lambda c: '.SH' if c[0] == '6' else '.SZ'

# ── 样式: xlsx skill标准 ──
F_HDR   = Font(name='Arial', size=9, bold=True)
F_LABEL = Font(name='Arial', size=9)
F_INPUT = Font(name='Arial', size=9)   # 全黑(用户要求)
F_CALC  = Font(name='Arial', size=9, color='000000')   # 黑=公式
FILL_HDR = PatternFill('solid', fgColor='F2F2F2')      # 浅灰表头,不用深色重底
THIN = Side(style='thin', color='D9D9D9')
BOT  = Border(bottom=Side(style='thin', color='000000'))
NUMFMT = '0.0%;(0.0%);"-"'   # 真百分比格式,负数括号,零显示-
FILL_IN = PatternFill('solid', fgColor='FFFF00')       # 黄底=需用户输入的假设格(xlsx标准)
# A股习惯: 涨红跌绿+加粗。用条件格式(非写死字体)→改数据/改区间自动跟随
RULE_UP = CellIsRule(operator='greaterThan', formula=['0'],
                     font=Font(name='Arial', size=9, bold=True, color='C00000'))
RULE_DN = CellIsRule(operator='lessThan', formula=['0'],
                     font=Font(name='Arial', size=9, bold=True, color='008000'))

VOLFMT = '0.00;(0.00);"-"'   # 量比: 1.00=常量, >1放量, <1缩量
RULE_VOL_HI = CellIsRule(operator='greaterThanOrEqual', formula=['1.5'],
                         font=Font(name='Arial', size=9, bold=True, color='C00000'))
RULE_VOL_LO = CellIsRule(operator='lessThanOrEqual', formula=['0.7'],
                         font=Font(name='Arial', size=9, bold=True, color='008000'))


def add_volsig(ws, rng):
    """量比信号: >=1.5放量(红粗) / <=0.7缩量(绿粗)"""
    ws.conditional_formatting.add(rng, RULE_VOL_HI)
    ws.conditional_formatting.add(rng, RULE_VOL_LO)


def add_updown(ws, rng):
    """给区域挂红涨绿跌加粗条件格式"""
    ws.conditional_formatting.add(rng, RULE_UP)
    ws.conditional_formatting.add(rng, RULE_DN)


def fetch(days):
    end = dt.date.today()
    start = end - dt.timedelta(days=int(days * 1.6) + 130)   # +130天:量比需60交易日历史打底
    allc = sorted({c for v in TREES.values() for c in v})
    data = {}; amt = {}
    for c in allc:
        h = ifd.history(c + sfx(c), start.isoformat(), end.isoformat(), 'close,amount')
        try:
            tb = h['tables'][0]
            data[c] = dict(zip(tb['time'], tb['table']['close']))
            amt[c] = dict(zip(tb['time'], tb['table']['amount']))
        except Exception:
            data[c] = {}; amt[c] = {}
    rt = ifd.realtime([c + sfx(c) for c in allc], 'latest,amount')
    today = end.isoformat()
    for c in allc:
        d = rt.get(c + sfx(c), {})
        if d.get('latest') and data.get(c):
            data[c][today] = d['latest']
        if d.get('amount') and amt.get(c) is not None:
            amt[c][today] = d['amount']
    return data, amt


def tree_daily(data, dates):
    out = {}
    for tree, codes in TREES.items():
        row = []
        for i in range(1, len(dates)):
            chg = [(data[c][dates[i]] / data[c][dates[i-1]] - 1) * 100
                   for c in codes
                   if data.get(c, {}).get(dates[i]) and data.get(c, {}).get(dates[i-1])]
            row.append(round(sum(chg) / len(chg) / 100, 6) if chg else None)
        out[tree] = row
    return out


def tree_volratio(amt, all_dates, want_dates, lookback=60):
    """每棵树的日度量比 = 当日成交额 / 过去lookback日均成交额
    ⛔用60日均而非5日均: 短均值会被恐慌天量污染(2026-08-04实证:7月三轮天量暴跌把5日均抬高,
    导致今日实际放量被误判为缩量)。60日能穿透单轮暴跌期。
    ⛔必须传all_dates(全量历史)算, 再按want_dates切片 —— 否则窗口开头几十天凑不满lookback会返空。"""
    idx = {d: i for i, d in enumerate(all_dates)}
    out = {}
    for tree, codes in TREES.items():
        full = {}
        for i, d in enumerate(all_dates):
            rs = []
            for c in codes:
                a = amt.get(c, {})
                cur = a.get(d)
                past = [a.get(x) for x in all_dates[max(0, i - lookback):i] if a.get(x)]
                if cur and len(past) >= 10:
                    rs.append(cur / (sum(past) / len(past)))
            full[d] = round(sum(rs) / len(rs), 3) if rs else None
        out[tree] = [full.get(d) for d in want_dates[1:]]
    return out


def build(history_days=180):
    os.makedirs(OUT_DIR, exist_ok=True)
    data, amt = fetch(history_days)
    all_dates = sorted({d for v in data.values() for d in v})
    dates_hist = all_dates[-(history_days + 1):]     # 历史sheet用
    dates30 = all_dates[-31:]                        # 近30日sheet用
    daily30 = tree_daily(data, dates30)
    daily_hist = tree_daily(data, dates_hist)
    vol30 = tree_volratio(amt, all_dates, dates30)
    vol_hist = tree_volratio(amt, all_dates, dates_hist)

    wb = load_workbook(OUT) if os.path.exists(OUT) else Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ══ Sheet1 近30日 (每天全刷新) ══
    if '近30日' in wb.sheetnames:
        wb.remove(wb['近30日'])
    ws = wb.create_sheet('近30日', 0)
    ndays = len(dates30) - 1                          # 30个日度列
    first_d = 5                                       # E列起是日期数据
    hdr = ['产品树', '近30日', '近10日', '近3日'] + [d[5:] for d in dates30[1:]]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(1, j, h)
        c.font = F_HDR; c.fill = FILL_HDR; c.border = BOT
        c.alignment = Alignment(horizontal='center', vertical='center')

    for i, (tree, codes) in enumerate(TREES.items(), 2):
        ws.cell(i, 1, tree).font = F_LABEL
        # 日度数据(蓝=硬编码输入)
        for k, v in enumerate(daily30[tree]):
            cc = ws.cell(i, first_d + k, v)
            cc.font = F_INPUT; cc.number_format = NUMFMT
            cc.alignment = Alignment(horizontal='center')
        # ⭐累计列用Excel公式(PRODUCT复利),黑色=公式
        last_col = get_column_letter(first_d + ndays - 1)
        for j, n in [(2, 30), (3, 10), (4, 3)]:
            if ndays >= n:
                s = get_column_letter(first_d + ndays - n)
                cols = [get_column_letter(first_d + ndays - n + k) for k in range(n)]
                f = '=' + '*'.join('(1+%s%d)' % (cl, i) for cl in cols) + '-1'
                cc = ws.cell(i, j, f)
                cc.font = F_CALC; cc.number_format = NUMFMT
                cc.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 22
    for j in range(2, 5):
        ws.column_dimensions[get_column_letter(j)].width = 9
    for j in range(first_d, first_d + ndays):
        ws.column_dimensions[get_column_letter(j)].width = 7
    ws.freeze_panes = 'E2'
    ws.sheet_view.showGridLines = False
    add_updown(ws, 'B2:D%d' % (len(TREES) + 1))          # 累计列红涨绿跌加粗

    # ══ Sheet2 历史累计 (只追加,不覆盖) ══
    HDR_ROW = 4          # 列头行(前3行是自选区间计算器)
    if '历史累计' not in wb.sheetnames:
        ws2 = wb.create_sheet('历史累计')
        # ── 第1-2行: 自选区间输入(黄格) ──
        c = ws2.cell(1, 1, '起始日'); c.font = F_HDR
        c = ws2.cell(2, 1, '截止日'); c.font = F_HDR
        for r, dv in [(1, '=A5'), (2, '=A%d' % (HDR_ROW + len(dates_hist)))]:
            cc = ws2.cell(r, 2, dv)
            cc.fill = FILL_IN; cc.font = Font(name='Arial', size=9, bold=True)
            cc.alignment = Alignment(horizontal='center')
            cc.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
        c = ws2.cell(1, 3, '← 在黄格填日期(YYYY-MM-DD),第3行累计自动重算'); c.font = F_LABEL
        # ── 第3行: 自选区间累计(公式,随黄格变) ──
        c = ws2.cell(3, 1, '区间累计'); c.font = F_HDR; c.fill = FILL_HDR
        # ── 第4行: 列头 ──
        c = ws2.cell(HDR_ROW, 1, '日期'); c.font = F_HDR; c.fill = FILL_HDR; c.border = BOT
        for j, tree in enumerate(TREES, 2):
            c = ws2.cell(HDR_ROW, j, tree)
            c.font = F_HDR; c.fill = FILL_HDR; c.border = BOT
            c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
        ws2.column_dimensions['A'].width = 12
        for j in range(2, len(TREES) + 2):
            ws2.column_dimensions[get_column_letter(j)].width = 11
        ws2.freeze_panes = 'B5'
        ws2.row_dimensions[HDR_ROW].height = 30
        ws2.sheet_view.showGridLines = False
    else:
        ws2 = wb['历史累计']

    existing = {ws2.cell(r, 1).value for r in range(HDR_ROW + 1, ws2.max_row + 1)}
    added = 0
    for k in range(1, len(dates_hist)):
        d = dates_hist[k]
        if d in existing:
            continue
        r = ws2.max_row + 1 if ws2.max_row > HDR_ROW else HDR_ROW + 1
        c = ws2.cell(r, 1, d)
        c.font = F_LABEL; c.alignment = Alignment(horizontal='center')
        for j, tree in enumerate(TREES, 2):
            cc = ws2.cell(r, j, daily_hist[tree][k-1])
            cc.font = F_INPUT; cc.number_format = NUMFMT
            cc.alignment = Alignment(horizontal='center')
        added += 1

    # ── 写第3行自选区间累计公式(随黄格B1/B2变化) ──
    last_r = ws2.max_row
    for j in range(2, len(TREES) + 2):
        col = get_column_letter(j)
        rng = '%s%d:%s%d' % (col, HDR_ROW + 1, col, last_r)
        drng = '$A$%d:$A$%d' % (HDR_ROW + 1, last_r)
        # 区间内(1+r)连乘 = EXP(SUMPRODUCT(条件*LN(1+r))) - 1
        # SUMPRODUCT天然按数组算,不需Ctrl+Shift+Enter;空格视为0,LN(1+0)=0不影响连乘
        f = ('=IFERROR(EXP(SUMPRODUCT(({d}>=$B$1)*({d}<=$B$2)*LN(1+{r})))-1,"")'
             .format(d=drng, r=rng))
        cc = ws2.cell(3, j, f)
        cc.font = F_CALC; cc.number_format = NUMFMT
        cc.alignment = Alignment(horizontal='center')
        cc.fill = PatternFill('solid', fgColor='FFF2CC')   # 淡黄底=计算结果区
    add_updown(ws2, 'B3:%s3' % get_column_letter(len(TREES) + 1))   # 区间累计行红涨绿跌加粗

    # ══ Sheet3 近30日量比 (结构对齐Sheet1) ══
    if '近30日量比' in wb.sheetnames:
        wb.remove(wb['近30日量比'])
    ws3 = wb.create_sheet('近30日量比')
    hdr3 = ['产品树', '近30日均', '近10日均', '近3日均'] + [d[5:] for d in dates30[1:]]
    for j, h in enumerate(hdr3, 1):
        c = ws3.cell(1, j, h)
        c.font = F_HDR; c.fill = FILL_HDR; c.border = BOT
        c.alignment = Alignment(horizontal='center', vertical='center')
    for i, tree in enumerate(TREES, 2):
        ws3.cell(i, 1, tree).font = F_LABEL
        for k, v in enumerate(vol30[tree]):
            cc = ws3.cell(i, first_d + k, v)
            cc.font = F_INPUT; cc.number_format = VOLFMT
            cc.alignment = Alignment(horizontal='center')
        lastc = get_column_letter(first_d + ndays - 1)
        for j, n in [(2, 30), (3, 10), (4, 3)]:
            if ndays >= n:
                st = get_column_letter(first_d + ndays - n)
                cc = ws3.cell(i, j, '=IFERROR(AVERAGE(%s%d:%s%d),"")' % (st, i, lastc, i))
                cc.font = F_CALC; cc.number_format = VOLFMT
                cc.alignment = Alignment(horizontal='center')
    ws3.column_dimensions['A'].width = 22
    for j in range(2, 5):
        ws3.column_dimensions[get_column_letter(j)].width = 9
    for j in range(first_d, first_d + ndays):
        ws3.column_dimensions[get_column_letter(j)].width = 7
    ws3.freeze_panes = 'E2'
    ws3.sheet_view.showGridLines = False
    add_volsig(ws3, 'B2:%s%d' % (lastc, len(TREES) + 1))
    ws3.cell(len(TREES) + 3, 1, '量比=当日成交额/前60日均成交额。>=1.5放量(红) <=0.7缩量(绿) 1.0=常量').font = F_LABEL

    # ══ Sheet4 历史量比 (结构对齐Sheet2,只追加) ══
    if '历史量比' not in wb.sheetnames:
        ws4 = wb.create_sheet('历史量比')
        c = ws4.cell(1, 1, '起始日'); c.font = F_HDR
        c = ws4.cell(2, 1, '截止日'); c.font = F_HDR
        for r, dv in [(1, '=A5'), (2, '=A%d' % (HDR_ROW + len(dates_hist)))]:
            cc = ws4.cell(r, 2, dv)
            cc.fill = FILL_IN; cc.font = Font(name='Arial', size=9, bold=True)
            cc.alignment = Alignment(horizontal='center')
            cc.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
        ws4.cell(1, 3, '← 黄格填日期,第3行=该区间平均量比').font = F_LABEL
        c = ws4.cell(3, 1, '区间均量比'); c.font = F_HDR; c.fill = FILL_HDR
        c = ws4.cell(HDR_ROW, 1, '日期'); c.font = F_HDR; c.fill = FILL_HDR; c.border = BOT
        for j, tree in enumerate(TREES, 2):
            c = ws4.cell(HDR_ROW, j, tree)
            c.font = F_HDR; c.fill = FILL_HDR; c.border = BOT
            c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
        ws4.column_dimensions['A'].width = 12
        for j in range(2, len(TREES) + 2):
            ws4.column_dimensions[get_column_letter(j)].width = 11
        ws4.freeze_panes = 'B5'
        ws4.row_dimensions[HDR_ROW].height = 30
        ws4.sheet_view.showGridLines = False
    else:
        ws4 = wb['历史量比']
    ex4 = {ws4.cell(r, 1).value for r in range(HDR_ROW + 1, ws4.max_row + 1)}
    add4 = 0
    for k in range(1, len(dates_hist)):
        d = dates_hist[k]
        if d in ex4:
            continue
        r = ws4.max_row + 1 if ws4.max_row > HDR_ROW else HDR_ROW + 1
        cc = ws4.cell(r, 1, d); cc.font = F_LABEL
        cc.alignment = Alignment(horizontal='center')
        for j, tree in enumerate(TREES, 2):
            c2 = ws4.cell(r, j, vol_hist[tree][k-1])
            c2.font = F_INPUT; c2.number_format = VOLFMT
            c2.alignment = Alignment(horizontal='center')
        add4 += 1
    lr4 = ws4.max_row
    for j in range(2, len(TREES) + 2):
        col = get_column_letter(j)
        f = ('=IFERROR(AVERAGEIFS({c}${h}:{c}${l},$A${h}:$A${l},">="&$B$1,$A${h}:$A${l},"<="&$B$2),"")'
             .format(c=col, h=HDR_ROW + 1, l=lr4))
        cc = ws4.cell(3, j, f)
        cc.font = F_CALC; cc.number_format = VOLFMT
        cc.alignment = Alignment(horizontal='center')
        cc.fill = PatternFill('solid', fgColor='FFF2CC')
    add_volsig(ws4, 'B3:%s3' % get_column_letter(len(TREES) + 1))

    wb.save(OUT)
    return OUT, len(TREES), added, dates30[-1], ws2.max_row - HDR_ROW, add4


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--history-days', type=int, default=180, help='历史sheet深度(默认180=6个月)')
    a = ap.parse_args()
    path, ntree, nadd, last, nhist, nvol = build(a.history_days)
    print('✓ %s' % path)
    print('  近30日  : %d棵树已刷新 | 最新 %s | 累计列=Excel公式(可随数据更新)' % (ntree, last))
    print('  历史累计: 新追加%d行, 现共%d行(已有日期跳过,不覆盖)' % (nadd, nhist))
    print('  近30日量比: 已刷新 | 历史量比: 新追加%d行' % nvol)
