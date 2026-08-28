# -*- coding: utf-8 -*-
"""美股中报财报日涨跌全表 — 渲染器
严格遵循 ~/.claude/standards/excel_model_standard.md (2026-08-04定版,唯一权威)
  等线11pt / 表头无填充只加粗 / 零边框 / 7个数字格式键 / 黄=可改假设 / 红=只标问题 / 能算的用公式
"""
import json, warnings
import pandas as pd, numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
warnings.filterwarnings('ignore')

# ── 标准 §4 字体 / §5 数字格式(仅7键) / §2 配色 ──────────────────
FONT = "等线"
def F(b=False, sz=11, color="000000"): return Font(name=FONT, size=sz, bold=b, color=color)
FMT = {
    'accounting_int': '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)',
    'accounting_1dp': '_(* #,##0.0_);_(* \\(#,##0.0\\);_(* "-"??_);_(@_)',
    'accounting_2dp': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)',
    'pct':  '0%', 'pct1': '0.0%',
    'usd':  '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)',
    'text': '@',
}
YEL = PatternFill("solid", fgColor="FFFF00")   # §2 只给可改假设
RED = PatternFill("solid", fgColor="FFCCCC")   # §2 只给四类问题
R = Alignment(horizontal="right"); L = Alignment(horizontal="left")

def put(ws, r, c, v, key='text', bold=False, fill=None, note=None, align=None):
    x = ws.cell(r, c, v); x.font = F(b=bold); x.number_format = FMT[key]
    if fill: x.fill = fill
    if note: x.comment = Comment(note, "Claude", height=90, width=300)
    x.alignment = align if align else (R if key != 'text' else L)
    return x

def widths(ws, spec):
    for col, w in spec.items(): ws.column_dimensions[col].width = w

def sheetfont(ws):
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None and (c.font is None or c.font.name != FONT):
                c.font = F(b=bool(c.font and c.font.bold))

# ── 数据装配 ─────────────────────────────────────────────────────
def build_data():
    rows = json.load(open('/tmp/earn_rows.json'))
    px   = pd.read_pickle('/tmp/earn_update/px.pkl')
    meta = json.load(open('/tmp/earn_update/meta.json'))
    new  = json.load(open('/tmp/earn_update/new_events.json'))
    last_day = px.index[-1]

    def reaction_row(tk, edate, name=None, tagA=None, tagB=None, session='盘前'):
        """由财报日+时点算 反应日/前日收盘/反应日收盘/现价
        ⛔ 盘前(<12点ET)=当日即反应; 盘后(>=12点ET)=次一交易日才反应(N-03)"""
        if tk not in px.columns: return None
        s = px[tk].dropna()
        if s.empty: return None
        d = pd.Timestamp(edate).normalize()
        after = s.index[s.index > d] if session == '盘后' else s.index[s.index >= d]
        if len(after) == 0: return None
        rday = after[0]
        prior = s.index[s.index < rday]
        if len(prior) == 0: return None
        pday = prior[-1]
        return dict(代码=tk, 公司=name or meta.get(tk,{}).get('name',tk),
                    板块A=tagA, 板块B=tagB, 财报日=str(d.date()), 反应日=str(rday.date()),
                    前日收盘=float(s.loc[pday]), 反应日收盘=float(s.loc[rday]),
                    现价=float(s.iloc[-1]), h52=meta.get(tk,{}).get('h52'))

    out, seen = [], set()
    for r in rows:                                   # 存量510: 保留原板块标签,价格全刷新
        tk = r['代码']
        rec = reaction_row(tk, r['财报日'], r['公司'], r['板块Tag 新闻口径'], r['板块Tag 扩展'], r['时点'])
        if rec: rec['时点'] = r['时点']; out.append(rec); seen.add(tk)
    ind2tag = {}                                     # 新增股按 industry 映射到扩展Tag
    for r in rows:
        t = meta.get(r['代码'],{}).get('industry')
        if t and r['板块Tag 扩展']: ind2tag.setdefault(t, r['板块Tag 扩展'])
    for e in new:                                    # 新增(8/21-8/27 那批,含NVDA)
        tk = e['t']
        if tk in seen: continue
        tag = ind2tag.get(meta.get(tk,{}).get('industry')) or '📌 其他'
        ts = pd.Timestamp(e['dt']); sess = '盘后' if ts.hour >= 12 else '盘前'
        rec = reaction_row(tk, ts, None, '📌 其他', tag, sess)
        if rec: rec['时点'] = sess; rec['是新增'] = True; out.append(rec)

    for r in out:                                    # 派生
        r['当日涨跌'] = r['反应日收盘']/r['前日收盘'] - 1
        r['至今']     = r['现价']/r['反应日收盘'] - 1
        r['简单和']   = r['当日涨跌'] + r['至今']
        r['全程']     = r['现价']/r['前日收盘'] - 1
        r['距52高']   = (r['现价']/r['h52'] - 1) if r.get('h52') else None
    out.sort(key=lambda x: -x['当日涨跌'])
    for i, r in enumerate(out, 1): r['排名'] = i
    return out, str(last_day.date())

# ── Sheet 01 读法与假设(Cover) ────────────────────────────────────
def sheet_cover(wb, data, asof, stats):
    ws = wb.create_sheet('01_读法与假设')
    widths(ws, {'A':34,'B':13,'C':13,'D':13,'E':13,'F':13,'G':13,'H':13})
    r = 1
    put(ws,r,1,'美股 2026 中报 · 财报日涨跌全表',bold=True); r+=1
    put(ws,r,1,f'数据截至 {asof} 收盘 · 全样本 {len(data)} 只'); r+=2

    put(ws,r,1,'可改假设',bold=True); r+=1
    put(ws,r,1,'结局阈值 上沿')
    c=put(ws,r,2,0.03,'pct1',fill=YEL,note='反应日至今 高于此值 = 守住并续涨。改这一格,全表结局列重算。')
    wb.defined_names.add(DefinedName('THRESH_UP', attr_text=f"'01_读法与假设'!$B${r}")); r+=1
    put(ws,r,1,'结局阈值 下沿')
    put(ws,r,2,-0.03,'pct1',fill=YEL,note='反应日至今 低于此值 = 回吐。改这一格,全表结局列重算。')
    wb.defined_names.add(DefinedName('THRESH_DN', attr_text=f"'01_读法与假设'!$B${r}")); r+=1
    put(ws,r,1,'极端涨跌门槛')
    put(ws,r,2,0.10,'pct',fill=YEL,note='用于切片表的 涨>N / 跌>N 家数统计。')
    wb.defined_names.add(DefinedName('EXTREME', attr_text=f"'01_读法与假设'!$B${r}")); r+=2

    put(ws,r,1,'本季五条事实',bold=True); r+=1
    for k,v,key,note in stats:
        put(ws,r,1,k); put(ws,r,2,v,key,note=note); r+=1
    r+=1
    put(ws,r,1,'怎么读这张表',bold=True); r+=1
    for t in ['① 主表按当日涨跌降序。当日涨跌 = 反应日收盘 ÷ 前一交易日收盘 - 1。',
              '② 不看卖方 consensus surprise。价格本身是唯一的 surprise,涨了就是超了(Buwen 口径 N-04)。',
              '③ 全程% = 现价 ÷ 财报前日收盘 - 1,是真实的从财报前持有至今总回报(本版新增,优于旧的算术和)。',
              '④ 结局三分类由 THRESH_UP / THRESH_DN 两个黄格子驱动,改它们全表重算。',
              '⑤ 20_切片分析 是七个区块的矩阵:板块做列、指标做行,可直接选区画图。',
              '⑥ 每格的口径疑问看最右 N- 编号,回 91 表查。']:
        put(ws,r,1,t); r+=1
    sheetfont(ws); return ws

# ── Sheet 10 主表(大事实表: 行=记录, 标准§1.9 例外) ──────────────
MAIN_COLS = ['排名','代码','公司','板块Tag 新闻口径','板块Tag 扩展','财报日','反应日','时点',
             '前日收盘$','反应日收盘$','当日涨跌%','现价$','反应日至今%','全程%','简单和%','结局','距52周高%','口径']
def sheet_main(wb, data, asof=None):
    ws = wb.create_sheet('10_财报日涨跌全表')
    widths(ws, {'A':6,'B':8,'C':30,'D':22,'E':22,'F':11,'G':11,'H':7,'I':12,'J':13,
                'K':11,'L':11,'M':13,'N':11,'O':11,'P':12,'Q':12,'R':7})
    for i,t in enumerate(MAIN_COLS,1):
        c = ws.cell(1,i,t); c.font = F(b=True); c.alignment = R if i>=9 else L   # §2 表头无填充只加粗
    for n,d in enumerate(data):
        r = n+2
        put(ws,r,1,d['排名'],'accounting_int'); put(ws,r,2,d['代码']); put(ws,r,3,d['公司'])
        put(ws,r,4,d['板块A']); put(ws,r,5,d['板块B']); put(ws,r,6,d['财报日']); put(ws,r,7,d['反应日'])
        put(ws,r,8,d['时点'])
        put(ws,r,9,d['前日收盘'],'usd'); put(ws,r,10,d['反应日收盘'],'usd')
        put(ws,r,11,f'=J{r}/I{r}-1','pct1')                                   # §6 能算的用公式
        put(ws,r,12,d['现价'],'usd')
        put(ws,r,13,f'=L{r}/J{r}-1','pct1'); put(ws,r,14,f'=L{r}/I{r}-1','pct1')
        put(ws,r,15,f'=K{r}+M{r}','pct1')
        put(ws,r,16,f'=IF(G{r}="{asof}","待观察",IF(M{r}>THRESH_UP,"守住并续涨",IF(M{r}<THRESH_DN,"回吐","基本持平")))')
        put(ws,r,17,(f'=L{r}/{d["h52"]}-1' if d.get('h52') else None),'pct1')
        put(ws,r,18,'N-03' if not d.get('是新增') else 'N-13')
    ws.freeze_panes = 'C2'; ws.auto_filter.ref = f'A1:R{len(data)+1}'
    return ws

# ── Sheet 20 切片分析(密度铁律: 实体做列 指标做行, 七区块纵向铺开) ──
def _agg(sub, asof=None):
    """一组标的 → 17个指标。⛔反应日=最新交易日的(尚无观察窗口)不进结局统计"""
    if not sub: return [None]*17
    sub_e = [x for x in sub if x['反应日'] != asof] or sub
    d0=[x['当日涨跌'] for x in sub]; d1=[x['至今'] for x in sub]; d2=[x['全程'] for x in sub]
    best=max(sub,key=lambda x:x['当日涨跌']); worst=min(sub,key=lambda x:x['当日涨跌'])
    dn=[x for x in sub if x['当日涨跌']<0 and x['至今']<0]
    d1e=[x['至今'] for x in sub_e]
    return [len(sub), float(np.median(d0)), float(np.mean(d0)),
            sum(1 for v in d0 if v>0)/len(d0), sum(1 for v in d0 if v>=.10), sum(1 for v in d0 if v<=-.10),
            float(np.median(d1e)), sum(1 for v in d1e if v>0)/len(d1e), float(np.median(d2)),
            sum(1 for x in sub_e if x['至今']>.03), sum(1 for x in sub_e if -.03<=x['至今']<=.03),
            sum(1 for x in sub_e if x['至今']<-.03),
            sum(1 for x in sub_e if x['至今']<-.03)/len(sub_e),
            best['代码'], best['当日涨跌'], worst['代码'], worst['当日涨跌']]
IDX = [('家数','accounting_int'),('当日涨跌 中位','pct1'),('当日涨跌 均值','pct1'),('当日上涨占比','pct'),
       ('当日涨>10% 家数','accounting_int'),('当日跌>10% 家数','accounting_int'),
       ('反应日至今 中位','pct1'),('至今上涨占比','pct'),('全程 中位','pct1'),
       ('结局 守住并续涨','accounting_int'),('结局 基本持平','accounting_int'),('结局 回吐','accounting_int'),
       ('回吐占比','pct'),('当日最佳','text'),('最佳 当日%','pct1'),('当日最差','text'),('最差 当日%','pct1')]

def _matrix(ws, r, title, groups, note=None, asof=None):
    """一个区块: 分组做列, 17指标做行"""
    put(ws,r,1,title,bold=True,note=note); r+=1
    keys=list(groups.keys())
    for j,k in enumerate(keys):
        c=ws.cell(r,2+j,k); c.font=F(b=True); c.alignment=R
    r+=1
    vals={k:_agg(v,asof) for k,v in groups.items()}
    for i,(nm,key) in enumerate(IDX):
        put(ws,r,1,nm)
        for j,k in enumerate(keys):
            v=vals[k][i]
            put(ws,r,2+j,v,key if v is not None else 'text')
        r+=1
    return r+1

def sheet_slice(wb, data, asof=None):
    ws = wb.create_sheet('20_切片分析')
    widths(ws, {'A':22, **{get_column_letter(i):13 for i in range(2,24)}})
    r = 1
    put(ws,r,1,'切片分析 — 板块/分档/时点/极值',bold=True); r+=2

    byB={}
    for d in data: byB.setdefault(d['板块B'] or '📌 其他',[]).append(d)
    byB={k:v for k,v in sorted(byB.items(),key=lambda kv:-len(kv[1]))}
    r=_matrix(ws,r,'区块A 扩展板块Tag × 指标',byB,asof=asof,note='Claude 按 yfinance industry 映射的20个扩展桶(N-08),非 Buwen 原分类。')

    byA={}
    for d in data: byA.setdefault(d['板块A'] or '📌 其他',[]).append(d)
    byA={k:v for k,v in sorted(byA.items(),key=lambda kv:-len(kv[1]))}
    r=_matrix(ws,r,'区块B 新闻口径板块Tag × 指标',byA,asof=asof,note='Buwen 在 telegram 新闻群使用的原始 tag 口径(N-07)。')

    bands=[('跌>10%',lambda v:v<=-.10),('跌5-10%',lambda v:-.10<v<=-.05),('跌2-5%',lambda v:-.05<v<=-.02),
           ('平 ±2%',lambda v:-.02<v<.02),('涨2-5%',lambda v:.02<=v<.05),('涨5-10%',lambda v:.05<=v<.10),
           ('涨>10%',lambda v:v>=.10)]
    byBand={nm:[d for d in data if f(d['当日涨跌'])] for nm,f in bands}
    r=_matrix(ws,r,'区块C 当日涨幅分档 × 后续表现',byBand,asof=asof,note='检验"追财报暴涨"是否成立: 看 涨>10% 那一列的 反应日至今中位 与 回吐占比。')

    byT={}
    for d in data: byT.setdefault(d['时点'],[]).append(d)
    byT['全样本']=data
    r=_matrix(ws,r,'区块D 财报时点 × 指标',byT,asof=asof,note='盘前=当日即反应; 盘后=次一交易日反应(N-03)。')

    put(ws,r,1,'区块E 当日涨幅榜 前30',bold=True); r+=1
    hdr=['排名','代码','公司','板块Tag 扩展','财报日','当日涨跌%','反应日至今%','全程%','结局']
    for j,h in enumerate(hdr):
        c=ws.cell(r,1+j,h); c.font=F(b=True); c.alignment=R if j>=5 else L
    r+=1
    for d in data[:30]:
        put(ws,r,1,d['排名'],'accounting_int'); put(ws,r,2,d['代码']); put(ws,r,3,d['公司'][:28])
        put(ws,r,4,d['板块B']); put(ws,r,5,d['财报日'])
        put(ws,r,6,d['当日涨跌'],'pct1'); put(ws,r,7,d['至今'],'pct1'); put(ws,r,8,d['全程'],'pct1')
        put(ws,r,9,'待观察' if d['反应日']==asof else ('守住并续涨' if d['至今']>.03 else ('回吐' if d['至今']<-.03 else '基本持平'))); r+=1
    r+=1
    put(ws,r,1,'区块F 当日跌幅榜 前30',bold=True); r+=1
    for j,h in enumerate(hdr):
        c=ws.cell(r,1+j,h); c.font=F(b=True); c.alignment=R if j>=5 else L
    r+=1
    for d in sorted(data,key=lambda x:x['当日涨跌'])[:30]:
        put(ws,r,1,d['排名'],'accounting_int'); put(ws,r,2,d['代码']); put(ws,r,3,d['公司'][:28])
        put(ws,r,4,d['板块B']); put(ws,r,5,d['财报日'])
        put(ws,r,6,d['当日涨跌'],'pct1'); put(ws,r,7,d['至今'],'pct1'); put(ws,r,8,d['全程'],'pct1')
        put(ws,r,9,'待观察' if d['反应日']==asof else ('守住并续涨' if d['至今']>.03 else ('回吐' if d['至今']<-.03 else '基本持平'))); r+=1
    r+=1
    dbl=sorted([d for d in data if d['当日涨跌']<0 and d['至今']<0],key=lambda x:x['全程'])
    put(ws,r,1,f'区块G 财报双负筛 — 当日跌且之后继续跌 ({len(dbl)}只)',bold=True,
        note='双负 = 市场当天否定 + 之后继续否定。本季两个最大亏损 HUBB/HWM 都出自此形态,已固化为每日筛。'); r+=1
    for j,h in enumerate(hdr):
        c=ws.cell(r,1+j,h); c.font=F(b=True); c.alignment=R if j>=5 else L
    r+=1
    for d in dbl[:40]:
        put(ws,r,1,d['排名'],'accounting_int'); put(ws,r,2,d['代码']); put(ws,r,3,d['公司'][:28])
        put(ws,r,4,d['板块B']); put(ws,r,5,d['财报日'])
        put(ws,r,6,d['当日涨跌'],'pct1'); put(ws,r,7,d['至今'],'pct1'); put(ws,r,8,d['全程'],'pct1')
        put(ws,r,9,'回吐'); r+=1
    ws.freeze_panes='B3'
    return ws, len(dbl)

# ── Sheet 91 口径与来源(标准§7 专表) ──────────────────────────────
NOTES = [
 ('N-01','10_全表','A:R','样本','标普500 + Buwen 70领域扫描成分股 + 内部观察池,去重599只候选; 窗口内有财报且可算反应日的进表。'),
 ('N-02','10_全表','F','窗口','2026-06-25 至 2026-08-27(2026年Q2/中报季)。⛔本版已把上版排除的 NVDA(8/26)等收进来,窗口右端从8/20延到8/27。'),
 ('N-03','10_全表','K','当日涨跌','反应日收盘 ÷ 前一交易日收盘 - 1。反应日由财报时间戳定: <12点ET=盘前(当日反应), >=12点ET=盘后(次一交易日反应)。'),
 ('N-04','10_全表','—','不看consensus','Buwen 口径(2026-08-24): 卖方 consensus 是卖方观点不进分析链; 价格本身是唯一的 surprise,涨了就是超了。本表无超预期%列。'),
 ('N-05','10_全表','I,J,L','数据源','yfinance: Ticker.earnings_dates 取财报时间戳, download(auto_adjust=True) 取复权收盘。现价=2026-08-27收盘。'),
 ('N-06','10_全表','—','已剔除','MNST 剔除: 8/11 执行1拆2,yfinance 在 8/06 出现凭空腰斩次日还原的脏数据,按此算出 +91.93% 是数据假象。已核查全部|反应|>8%的111只,窗口内有拆股的仅此1只。'),
 ('N-07','10_全表','D','A列板块','Buwen 在 telegram 新闻群使用的原始 tag 口径。'),
 ('N-08','10_全表','E','B列板块','Claude 按 A列同样命名风格补的扩展桶,依据 yfinance industry 字段映射。非 Buwen 原分类。'),
 ('N-09','10_全表','N','全程%','现价 ÷ 财报前日收盘 - 1 = 从财报前一天持有至今的真实总回报。⭐本版新增,取代旧版"简单和"作为首选口径。'),
 ('N-10','10_全表','O','简单和%','当日涨跌% + 反应日至今%,算术相加。上版的口径,保留仅为跨版本可比; ⛔它不是真实收益(真实收益看 N-09)。'),
 ('N-11','10_全表','P','结局判定','由 01表 两个黄格子 THRESH_UP(+3%)/THRESH_DN(-3%) 驱动。改黄格子,全表 510+ 行结局列重算。'),
 ('N-12','20_切片','区块C','分档口径','按当日涨跌分7档。用于检验"追财报暴涨"是否成立。'),
 ('N-13','10_全表','H','本版新增','时点标"新增"的32只 = 上版(8/24)之后才发财报的(NVDA/CRM/CRWD/GEV等) + 上版 yfinance 漏抓补回的(GOOGL/GS/ICE等)。其反应日已按 N-03 同口径算。'),
 ('N-14','10_全表','P','待观察',"反应日 = 数据截止日(最新交易日)的标的,尚无'之后'的观察窗口,结局标'待观察'而非'基本持平'——其 反应日至今%必然为0,不代表市场没反应。切片表的结局类统计已剔除这批。"),
 ('N-15','全簿','—','格式标准','遵循 ~/.claude/standards/excel_model_standard.md(2026-08-04定版): 等线11pt / 表头无填充只加粗 / 零边框 / 7个数字格式键 / 黄=可改假设 / 红=只标问题 / 能算的用公式。'),
]
def sheet_notes(wb):
    ws = wb.create_sheet('91_口径与来源')
    widths(ws, {'A':8,'B':13,'C':10,'D':16,'E':110})
    for j,h in enumerate(['编号','所属表','单元格','项目','口径 / 来源 / 判断']):
        c=ws.cell(1,1+j,h); c.font=F(b=True)
    for i,row in enumerate(NOTES,2):
        for j,v in enumerate(row): put(ws,i,1+j,v)
        ws.row_dimensions[i].height = 15
    ws.freeze_panes='A2'; ws.auto_filter.ref=f'A1:E{len(NOTES)+1}'
    return ws

def main():
    data, asof = build_data()
    d0=[x['当日涨跌'] for x in data]; d1=[x['至今'] for x in data]
    corr=float(np.corrcoef(d0,d1)[0,1])
    up10=[x for x in data if x['当日涨跌']>=.10]
    stats=[('全样本家数',len(data),'accounting_int','窗口内有财报且可算反应日的全部标的。'),
           ('财报日涨跌 中位',float(np.median(d0)),'pct1','中位数个股在自己财报日的表现。'),
           ('财报日上涨占比',sum(1 for v in d0 if v>0)/len(d0),'pct','不到一半=财报日整体不是利好事件。'),
           ('当日vs之后 相关系数',corr,'accounting_2dp','负值=均值回归: 当天涨得多的之后反而弱。'),
           ('涨>10%后仍守住占比',(sum(1 for x in up10 if x["至今"]>.03)/len(up10)) if up10 else 0,'pct','追财报暴涨的实际胜率。')]
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    sheet_cover(wb, data, asof, stats)
    sheet_main(wb, data, asof)
    ws20, ndbl = sheet_slice(wb, data, asof)
    sheet_notes(wb)
    for ws in wb: sheetfont(ws)
    out='/Users/huaichuaibeimeng/Desktop/美股2026中报_财报日涨跌全表.xlsx'
    wb.save(out)
    print(f"✓ 保存 {out}")
    print(f"  全样本 {len(data)} 只 (上版510, 新增{sum(1 for d in data if d.get('是新增'))}) | 截至 {asof}")
    print(f"  中位 {np.median(d0):+.2%} | 上涨占比 {sum(1 for v in d0 if v>0)/len(d0):.1%} | 相关 {corr:+.3f} | 双负 {ndbl}只")
    return out

if __name__ == '__main__':
    main()
